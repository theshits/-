from fastapi import APIRouter, Depends, HTTPException, File, UploadFile
from sqlalchemy.orm import Session
from sqlalchemy.exc import SQLAlchemyError
from typing import List, Optional
import logging
from backend.database import get_db
from backend.models.models import EmissionSource, PollutantEmission, POLLUTANT_TYPES, MARKER_SYMBOLS
from backend.models.schemas import (
    EmissionSourceCreate, 
    EmissionSourceUpdate, 
    EmissionSourceResponse,
    PollutantEmissionResponse,
    PollutantTypeInfo,
    MarkerSymbolInfo
)

router = APIRouter()
logger = logging.getLogger(__name__)

@router.get("/", response_model=List[EmissionSourceResponse])
def get_sources(
    skip: int = 0, 
    limit: int = 100, 
    db: Session = Depends(get_db)
):
    sources = db.query(EmissionSource).offset(skip).limit(limit).all()
    return sources

@router.get("/pollutant-types/", response_model=List[PollutantTypeInfo])
def get_pollutant_types():
    return [
        PollutantTypeInfo(
            type=k,
            name=v['name'],
            unit=v['unit'],
            description=v['description']
        )
        for k, v in POLLUTANT_TYPES.items()
    ]

@router.get("/marker-symbols/", response_model=List[MarkerSymbolInfo])
def get_marker_symbols():
    return [
        MarkerSymbolInfo(
            symbol=k,
            name=v['name'],
            icon=v['icon']
        )
        for k, v in MARKER_SYMBOLS.items()
    ]

@router.post("/", response_model=EmissionSourceResponse)
def create_source(source: EmissionSourceCreate, db: Session = Depends(get_db)):
    try:
        source_data = source.model_dump(exclude={'pollutants'})
        db_source = EmissionSource(**source_data)
        db.add(db_source)
        db.flush()
        
        for pollutant in source.pollutants:
            db_pollutant = PollutantEmission(
                source_id=db_source.id,
                pollutant_type=pollutant.pollutant_type,
                emission_rate=pollutant.emission_rate,
                concentration=pollutant.concentration
            )
            db.add(db_pollutant)
        
        db.commit()
        db.refresh(db_source)
        return db_source
    except SQLAlchemyError as e:
        db.rollback()
        logger.error(f"创建排放源失败: {e}")
        raise HTTPException(status_code=500, detail=f"数据库错误: {str(e)}")
    except Exception as e:
        db.rollback()
        logger.error(f"创建排放源失败: {e}")
        raise HTTPException(status_code=500, detail=f"创建失败: {str(e)}")

@router.post("/batch", response_model=List[EmissionSourceResponse])
def create_sources_batch(
    sources: List[EmissionSourceCreate], 
    db: Session = Depends(get_db)
):
    db_sources = []
    for source in sources:
        source_data = source.model_dump(exclude={'pollutants'})
        db_source = EmissionSource(**source_data)
        db.add(db_source)
        db.flush()
        
        for pollutant in source.pollutants:
            db_pollutant = PollutantEmission(
                source_id=db_source.id,
                pollutant_type=pollutant.pollutant_type,
                emission_rate=pollutant.emission_rate,
                concentration=pollutant.concentration
            )
            db.add(db_pollutant)
        
        db_sources.append(db_source)
    
    db.commit()
    for source in db_sources:
        db.refresh(source)
    return db_sources

@router.get("/{source_id}", response_model=EmissionSourceResponse)
def get_source(source_id: int, db: Session = Depends(get_db)):
    source = db.query(EmissionSource).filter(EmissionSource.id == source_id).first()
    if not source:
        raise HTTPException(status_code=404, detail="排放源未找到")
    return source

@router.put("/{source_id}", response_model=EmissionSourceResponse)
def update_source(
    source_id: int, 
    source: EmissionSourceUpdate, 
    db: Session = Depends(get_db)
):
    db_source = db.query(EmissionSource).filter(EmissionSource.id == source_id).first()
    if not db_source:
        raise HTTPException(status_code=404, detail="排放源未找到")
    
    update_data = source.model_dump(exclude_unset=True, exclude={'pollutants'})
    for key, value in update_data.items():
        setattr(db_source, key, value)
    
    if source.pollutants is not None:
        db.query(PollutantEmission).filter(PollutantEmission.source_id == source_id).delete()
        
        for pollutant in source.pollutants:
            db_pollutant = PollutantEmission(
                source_id=source_id,
                pollutant_type=pollutant.pollutant_type,
                emission_rate=pollutant.emission_rate,
                concentration=pollutant.concentration
            )
            db.add(db_pollutant)
    
    db.commit()
    db.refresh(db_source)
    return db_source

@router.delete("/{source_id}")
def delete_source(source_id: int, db: Session = Depends(get_db)):
    db_source = db.query(EmissionSource).filter(EmissionSource.id == source_id).first()
    if not db_source:
        raise HTTPException(status_code=404, detail="排放源未找到")
    
    db.delete(db_source)
    db.commit()
    return {"message": "排放源已删除", "id": source_id}

@router.post("/{source_id}/pollutants/", response_model=PollutantEmissionResponse)
def add_pollutant(
    source_id: int,
    pollutant_type: str,
    emission_rate: float = 0.0,
    concentration: Optional[float] = None,
    db: Session = Depends(get_db)
):
    db_source = db.query(EmissionSource).filter(EmissionSource.id == source_id).first()
    if not db_source:
        raise HTTPException(status_code=404, detail="排放源未找到")
    
    if pollutant_type not in POLLUTANT_TYPES:
        raise HTTPException(status_code=400, detail="无效的污染物类型")
    
    existing = db.query(PollutantEmission).filter(
        PollutantEmission.source_id == source_id,
        PollutantEmission.pollutant_type == pollutant_type
    ).first()
    
    if existing:
        existing.emission_rate = emission_rate
        existing.concentration = concentration
        db.commit()
        db.refresh(existing)
        return existing
    
    db_pollutant = PollutantEmission(
        source_id=source_id,
        pollutant_type=pollutant_type,
        emission_rate=emission_rate,
        concentration=concentration
    )
    db.add(db_pollutant)
    db.commit()
    db.refresh(db_pollutant)
    return db_pollutant

@router.delete("/{source_id}/pollutants/{pollutant_id}")
def delete_pollutant(
    source_id: int,
    pollutant_id: int,
    db: Session = Depends(get_db)
):
    db_pollutant = db.query(PollutantEmission).filter(
        PollutantEmission.id == pollutant_id,
        PollutantEmission.source_id == source_id
    ).first()
    
    if not db_pollutant:
        raise HTTPException(status_code=404, detail="污染物排放记录未找到")
    
    db.delete(db_pollutant)
    db.commit()
    return {"message": "污染物排放记录已删除", "id": pollutant_id}

@router.get("/template/{source_type}")
async def download_template(source_type: str):
    """下载排放源导入模板"""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    
    wb = Workbook()
    ws = wb.active
    
    pollutant_types = list(POLLUTANT_TYPES.keys())
    
    base_headers = {
        'point': ['名称', '纬度', '经度', '高度', '烟气温度(K)', '烟气速度', '烟囱直径'],
        'area': ['名称', '纬度', '经度', '面源长度', '面源宽度', '面源高度', '烟气温度(K)'],
        'equivalent_area': ['名称', '纬度', '经度', '面源长度', '面源宽度', '面源高度', '烟气温度(K)'],
        'line': ['名称', '起点纬度', '起点经度', '终点纬度', '终点经度', '线源宽度', '线源高度', '烟气温度(K)']
    }
    
    if source_type not in base_headers:
        raise HTTPException(status_code=400, detail="无效的排放源类型")
    
    headers = base_headers[source_type] + pollutant_types + ['标记符号', '标记颜色']
    
    header_fill = PatternFill(start_color="007AFF", end_color="007AFF", fill_type="solid")
    pollutant_fill = PatternFill(start_color="34C759", end_color="34C759", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        if header in pollutant_types:
            cell.fill = pollutant_fill
        else:
            cell.fill = header_fill
    
    example_data = {
        'point': ['示例点源', 39.9, 116.4, 50, 400, 15, 2],
        'area': ['示例面源', 39.9, 116.4, 100, 100, 10, 300],
        'equivalent_area': ['示例等效面源', 39.9, 116.4, 100, 100, 10, 300],
        'line': ['示例线源', 39.9, 116.4, 39.91, 116.41, 10, 5, 300]
    }
    
    example_pollutants = {pt: '' for pt in pollutant_types}
    if pollutant_types:
        example_pollutants[pollutant_types[0]] = 10.5 if source_type != 'equivalent_area' else 100
    
    row_data = example_data[source_type] + [example_pollutants[pt] for pt in pollutant_types] + ['factory', '#FF5722']
    
    for col, value in enumerate(row_data, 1):
        ws.cell(row=2, column=col, value=value)
    
    for col in range(1, len(headers) + 1):
        col_letter = chr(64 + col) if col <= 26 else chr(64 + (col - 1) // 26) + chr(65 + (col - 1) % 26)
        ws.column_dimensions[col_letter].width = 15
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={source_type}_template.xlsx"}
    )

@router.post("/import/{source_type}")
async def import_sources(
    source_type: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """批量导入排放源"""
    from openpyxl import load_workbook
    import io
    
    if source_type not in ['point', 'area', 'equivalent_area', 'line']:
        raise HTTPException(status_code=400, detail="无效的排放源类型")
    
    try:
        content = await file.read()
        wb = load_workbook(io.BytesIO(content))
        ws = wb.active
        
        headers = [cell.value for cell in ws[1]]
        pollutant_types = list(POLLUTANT_TYPES.keys())
        pollutant_cols = {}
        for idx, header in enumerate(headers):
            if header in pollutant_types:
                pollutant_cols[header] = idx
        
        imported_count = 0
        errors = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not row[0]:
                continue
            
            try:
                marker_symbol_val = str(row[-2]).strip() if row[-2] else 'factory'
                marker_color_val = str(row[-1]).strip() if row[-1] else '#FF5722'
                
                if source_type == 'point':
                    source = EmissionSource(
                        name=str(row[0]).strip(),
                        source_type='point',
                        latitude=float(row[1]) if row[1] else 0,
                        longitude=float(row[2]) if row[2] else 0,
                        height=float(row[3]) if row[3] else 0,
                        temperature=float(row[4]) if row[4] else 400,
                        velocity=float(row[5]) if row[5] else 15,
                        diameter=float(row[6]) if row[6] else 2,
                        marker_symbol=marker_symbol_val,
                        marker_color=marker_color_val,
                        is_active=True
                    )
                elif source_type == 'area':
                    source = EmissionSource(
                        name=str(row[0]).strip(),
                        source_type='area',
                        latitude=float(row[1]) if row[1] else 0,
                        longitude=float(row[2]) if row[2] else 0,
                        area_length=float(row[3]) if row[3] else 100,
                        area_width=float(row[4]) if row[4] else 100,
                        area_height=float(row[5]) if row[5] else 0,
                        area_temperature=float(row[6]) if row[6] else 300,
                        marker_symbol=marker_symbol_val,
                        marker_color=marker_color_val,
                        is_active=True
                    )
                elif source_type == 'equivalent_area':
                    source = EmissionSource(
                        name=str(row[0]).strip(),
                        source_type='equivalent_area',
                        latitude=float(row[1]) if row[1] else 0,
                        longitude=float(row[2]) if row[2] else 0,
                        area_length=float(row[3]) if row[3] else 100,
                        area_width=float(row[4]) if row[4] else 100,
                        area_height=float(row[5]) if row[5] else 0,
                        area_temperature=float(row[6]) if row[6] else 300,
                        marker_symbol=marker_symbol_val,
                        marker_color=marker_color_val,
                        is_active=True
                    )
                elif source_type == 'line':
                    source = EmissionSource(
                        name=str(row[0]).strip(),
                        source_type='line',
                        latitude=float(row[1]) if row[1] else 0,
                        longitude=float(row[2]) if row[2] else 0,
                        start_lat=float(row[1]) if row[1] else 0,
                        start_lon=float(row[2]) if row[2] else 0,
                        end_lat=float(row[3]) if row[3] else 0,
                        end_lon=float(row[4]) if row[4] else 0,
                        line_width=float(row[5]) if row[5] else 10,
                        line_height=float(row[6]) if row[6] else 0,
                        line_temperature=float(row[7]) if row[7] else 300,
                        marker_symbol=marker_symbol_val,
                        marker_color=marker_color_val,
                        is_active=True
                    )
                
                db.add(source)
                db.flush()
                
                for pollutant_type, col_idx in pollutant_cols.items():
                    if col_idx < len(row) and row[col_idx] is not None and row[col_idx] != '':
                        value = float(row[col_idx])
                        if value > 0:
                            pollutant = PollutantEmission(
                                source_id=source.id,
                                pollutant_type=pollutant_type,
                                emission_rate=value if source_type != 'equivalent_area' else 0,
                                concentration=value if source_type == 'equivalent_area' else None
                            )
                            db.add(pollutant)
                
                imported_count += 1
                
            except Exception as e:
                errors.append(f"第{row_idx}行: {str(e)}")
                continue
        
        db.commit()
        
        return {
            "message": f"成功导入 {imported_count} 条记录",
            "imported_count": imported_count,
            "errors": errors
        }
        
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"导入失败: {str(e)}")
