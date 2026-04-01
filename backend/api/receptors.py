from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from typing import List
from backend.database import get_db
from backend.models.models import Receptor
from backend.models.schemas import (
    ReceptorCreate, 
    ReceptorUpdate, 
    ReceptorResponse
)
from fastapi.responses import StreamingResponse
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

router = APIRouter()

@router.get("/", response_model=List[ReceptorResponse])
def get_receptors(
    skip: int = 0, 
    limit: int = 100, 
    db: Session = Depends(get_db)
):
    receptors = db.query(Receptor).offset(skip).limit(limit).all()
    return receptors

@router.get("/{receptor_id}", response_model=ReceptorResponse)
def get_receptor(receptor_id: int, db: Session = Depends(get_db)):
    receptor = db.query(Receptor).filter(Receptor.id == receptor_id).first()
    if not receptor:
        raise HTTPException(status_code=404, detail="受体点未找到")
    return receptor

@router.post("/", response_model=ReceptorResponse)
def create_receptor(receptor: ReceptorCreate, db: Session = Depends(get_db)):
    db_receptor = Receptor(**receptor.model_dump())
    db.add(db_receptor)
    db.commit()
    db.refresh(db_receptor)
    return db_receptor

@router.put("/{receptor_id}", response_model=ReceptorResponse)
def update_receptor(
    receptor_id: int, 
    receptor: ReceptorUpdate, 
    db: Session = Depends(get_db)
):
    db_receptor = db.query(Receptor).filter(Receptor.id == receptor_id).first()
    if not db_receptor:
        raise HTTPException(status_code=404, detail="受体点未找到")
    
    update_data = receptor.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(db_receptor, key, value)
    
    db.commit()
    db.refresh(db_receptor)
    return db_receptor

@router.delete("/{receptor_id}")
def delete_receptor(receptor_id: int, db: Session = Depends(get_db)):
    db_receptor = db.query(Receptor).filter(Receptor.id == receptor_id).first()
    if not db_receptor:
        raise HTTPException(status_code=404, detail="受体点未找到")
    
    db.delete(db_receptor)
    db.commit()
    return {"message": "受体点已删除", "id": receptor_id}

@router.post("/batch", response_model=List[ReceptorResponse])
def create_receptors_batch(
    receptors: List[ReceptorCreate], 
    db: Session = Depends(get_db)
):
    db_receptors = [Receptor(**receptor.model_dump()) for receptor in receptors]
    db.add_all(db_receptors)
    db.commit()
    for receptor in db_receptors:
        db.refresh(receptor)
    return db_receptors

@router.get("/template")
async def download_template():
    """下载受体点导入模板"""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    
    wb = Workbook()
    ws = wb.active
    
    headers = ['名称', '纬度', '经度', '高度', '标记符号', '标记颜色']
    
    header_fill = PatternFill(start_color="007AFF", end_color="007AFF", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.fill = header_fill
    
    example_data = ['示例受体点', 39.9, 116.4, 0, 'monitor', '#2196F3']
    
    for col, value in enumerate(example_data, 1):
        ws.cell(row=2, column=col, value=value)
    
    for col in range(1, len(headers) + 1):
        col_letter = chr(64 + col) if col <= 26 else chr(64 + (col - 1) % 26) + chr(65 + (col - 1) // 26)
        ws.column_dimensions[col_letter].width = 15
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=receptors_template.xlsx"}
    )

@router.post("/import")
async def import_receptors(
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """批量导入受体点"""
    from openpyxl import load_workbook
    import io
    
    try:
        content = await file.read()
        wb = load_workbook(io.BytesIO(content))
        ws = wb.active
        
        headers = [cell.value for cell in ws[1]]
        
        imported_count = 0
        errors = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not row[0]:
                continue
            
            try:
                marker_symbol_val = str(row[4]).strip() if row[4] else 'monitor'
                marker_color_val = str(row[5]).strip() if row[5] else '#2196F3'
                
                receptor = Receptor(
                    name=str(row[0]).strip(),
                    latitude=float(row[1]) if row[1] else 0,
                    longitude=float(row[2]) if row[2] else 0,
                    height=float(row[3]) if row[3] else 0,
                    marker_symbol=marker_symbol_val,
                    marker_color=marker_color_val,
                    is_active=True
                )
                
                db.add(receptor)
                db.flush()
                imported_count += 1
                
            except Exception as e:
                errors.append(f"第{row_idx}行: {str(e)}")
        
        db.commit()
        
        if errors:
            return {
                "imported_count": imported_count,
                "errors": errors,
                "message": f"导入完成：成功{imported_count}个，失败{len(errors)}个"
            }
        
        return {
            "imported_count": imported_count,
            "message": f"导入成功！共导入{imported_count}个受体点"
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"导入失败: {str(e)}")

@router.post("/export")
async def export_receptors(
    ids: List[int],
    db: Session = Depends(get_db)
):
    """批量导出受体点"""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    
    receptors = db.query(Receptor).filter(Receptor.id.in_(ids)).all()
    
    if not receptors:
        raise HTTPException(status_code=404, detail="未找到指定的受体点")
    
    headers = ['名称', '纬度', '经度', '高度', '标记符号', '标记颜色']
    
    wb = Workbook()
    ws = wb.active
    
    header_fill = PatternFill(start_color="007AFF", end_color="007AFF", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.fill = header_fill
    
    for receptor in receptors:
        row_data = [
            receptor.name,
            receptor.latitude,
            receptor.longitude,
            receptor.height,
            receptor.marker_symbol or 'monitor',
            receptor.marker_color or '#2196F3'
        ]
        
        for col, value in enumerate(row_data, 1):
            ws.cell(row=receptors.index(receptor) + 2, column=col, value=value)
    
    for col in range(1, len(headers) + 1):
        col_letter = chr(64 + col) if col <= 26 else chr(64 + (col - 1) % 26) + chr(65 + (col - 1) // 26)
        ws.column_dimensions[col_letter].width = 15
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=receptors_export.xlsx"}
    )
